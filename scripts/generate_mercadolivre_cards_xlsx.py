from __future__ import annotations

import argparse
import html
import re
import sys
import unicodedata
import webbrowser
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError as exc:  # pragma: no cover - runtime guard
    raise SystemExit(
        "Dependencia ausente: openpyxl. Instale com: ./.venv/bin/pip install openpyxl"
    ) from exc

from app.core.config import get_settings


CATEGORY_SUBCOLLECTION = "items"
HEADER_SCAN_MAX_ROWS = 12
HEADERS = [
    "titulo",
    "descricao",
    "preco",
    "quantidade",
    "sku",
    "foto_principal_url",
    "fotos_urls",
    "condicao",
    "idioma",
    "categoria_interna",
    "slug",
    "acabamento",
    "preco_base_origem",
]

TEMPLATE_HEADER_ALIASES: dict[str, set[str]] = {
    "titulo": {
        "titulo",
        "titulo do anuncio",
        "titulo do produto",
        "nome",
        "nome do produto",
    },
    "descricao": {
        "descricao",
        "descricao do produto",
        "descricao completa",
    },
    "preco": {
        "preco",
        "preco unitario",
        "valor",
        "valor unitario",
    },
    "quantidade": {
        "quantidade",
        "estoque",
        "qtd",
        "quantidade disponivel",
    },
    "sku": {
        "sku",
        "seller sku",
        "codigo sku",
        "id sku",
    },
    "foto_principal_url": {
        "foto principal",
        "imagem principal",
        "url da imagem",
        "imagem",
        "imagem 1",
        "foto 1",
        "fotos",
        "imagens",
    },
    "fotos_urls": {
        "fotos",
        "imagens",
        "urls de imagens",
        "url das imagens",
        "imagem 2",
        "foto 2",
    },
    "condicao": {
        "condicao",
        "estado",
        "tipo de condicao",
    },
    "idioma": {
        "idioma",
        "linguagem",
        "language",
    },
    "variacao_nome_carta": {
        "varia por nome do carta colecionavel",
        "nome do carta colecionavel",
        "nome da carta",
    },
    "codigo_universal_produto": {
        "codigo universal de produto",
        "gtin",
        "ean",
        "upc",
    },
    "motivo_gtin_vazio": {
        "motivo de gtin vazio",
        "motivo gtin vazio",
    },
    "unidade_venda": {
        "unidade de venda",
    },
    "tipo_anuncio": {
        "tipo de anuncio",
        "tipo do anuncio",
        "listing type",
    },
    "marca": {
        "marca",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Gera planilha XLSX para anuncio em massa no Mercado Livre "
            "somente com cartas (single_card) da Legacy Cards."
        )
    )
    parser.add_argument(
        "--database",
        default="legacy-cards",
        help="Database Firestore alvo. Padrao: legacy-cards",
    )
    parser.add_argument(
        "--output-dir",
        default="exports",
        help="Diretorio de saida para XLSX e HTML. Padrao: exports",
    )
    parser.add_argument(
        "--template-xlsx",
        default="",
        help=(
            "Caminho da planilha ORIGINAL baixada do Mercado Livre. "
            "Quando informado, o script preenche esse template mantendo estrutura."
        ),
    )
    parser.add_argument(
        "--open-html",
        action="store_true",
        help="Abre automaticamente o preview HTML no navegador.",
    )
    return parser.parse_args()


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _normalize_header(value: Any) -> str:
    text = _safe_str(value)
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _condition_tag(raw_condition: str) -> str:
    text = _normalize_key(raw_condition)
    if "nm" in text or "near mint" in text:
        return "NM"
    match = re.search(r"\(([a-z]{2,5})\)", text)
    if match:
        return match.group(1).upper()
    tokens = re.findall(r"[a-z]{2,5}", text)
    if tokens:
        return tokens[0].upper()
    return "NM"


def _truncate_title(value: str, limit: int = 60) -> str:
    safe = value.strip()
    if len(safe) <= limit:
        return safe
    return safe[:limit].rstrip()


def _finish_label_for_title(raw_finish: str) -> str:
    finish = _normalize_key(raw_finish)
    if "master ball" in finish:
        return "masterball"
    if "poke ball" in finish:
        return "pokeball"
    if "element" in finish and ("foil" in finish or "holo" in finish or "reverse" in finish):
        return "element foil"
    if "reverse" in finish and ("foil" in finish or "holo" in finish):
        return "reverse foil"
    if "mirror" in finish or "foil" in finish or "holo" in finish:
        return "foil"
    return "normal"


def _apply_price_rule(base_price_brl: float, raw_finish: str) -> tuple[float, str]:
    finish = _normalize_key(raw_finish)
    if "master ball" in finish:
        return 30.0, "master_ball_30"
    if "poke ball" in finish:
        return 20.0, "poke_ball_20"
    if "element" in finish and ("foil" in finish or "holo" in finish or "reverse" in finish):
        return 10.0, "element_foil_10"
    if "reverse" in finish and ("foil" in finish or "holo" in finish):
        return 10.0, "reverse_foil_10"
    if base_price_brl <= 0.50:
        return 1.0, "base_ate_50_centavos_1"
    return 2.0, "base_acima_50_centavos_2"


def _build_description(
    *,
    card_name: str,
    condition: str,
    finish: str,
    language: str,
    lot_code: str,
) -> str:
    condition_text = condition or "Near Mint (NM)"
    finish_text = finish or "Normal"
    language_text = language or "PT"
    normalized_lot = _normalize_key(lot_code)
    if normalized_lot in {"", "sem-lote", "sem lote", "n/a", "na", "none", "null", "-"}:
        safe_lot = "lote-1.1"
    else:
        safe_lot = lot_code
    return (
        f"Legacy Cards | Carta Pokémon original: {card_name}. "
        f"Condição: {condition_text}. Idioma: {language_text}. Acabamento: {finish_text}. "
        "As imagens são ilustrativas e representam a carta base. "
        "Nas versões foil/reverse foil, o brilho e efeito visual estão presentes na carta física. "
        "Produto separado e enviado pela Legacy Cards."
        f"\n\nCódigo: {safe_lot}"
    )


def _resolve_service_account_path() -> Path:
    settings = get_settings()
    configured = Path(settings.firestore_service_account_path)
    if configured.is_absolute():
        return configured
    cwd_path = Path.cwd() / configured
    if cwd_path.exists():
        return cwd_path
    return settings.backend_root / configured


def _get_firestore_client(database: str):
    from google.cloud import firestore
    from google.oauth2 import service_account

    settings = get_settings()
    project_id = settings.firestore_project_id
    service_account_path = _resolve_service_account_path()
    if not service_account_path.exists():
        raise RuntimeError(
            f"Service account nao encontrado em {service_account_path}. "
            "Ajuste FIRESTORE_SERVICE_ACCOUNT_PATH."
        )

    credentials = service_account.Credentials.from_service_account_file(str(service_account_path))
    resolved_project = project_id or credentials.project_id
    if not resolved_project:
        raise RuntimeError("Nao foi possivel resolver FIRESTORE_PROJECT_ID.")

    return firestore.Client(
        project=resolved_project,
        credentials=credentials,
        database=database,
    )


def fetch_cards_rows(database: str) -> tuple[list[dict[str, Any]], Counter]:
    settings = get_settings()
    client = _get_firestore_client(database)
    rows: list[dict[str, Any]] = []
    counters = Counter()

    for snapshot in client.collection_group(CATEGORY_SUBCOLLECTION).stream():
        payload = snapshot.to_dict() or {}
        owner_type = _normalize_key(_safe_str(payload.get("owner_type")) or "admin")
        product_type = _normalize_key(_safe_str(payload.get("product_type")))
        if owner_type != "admin" or product_type != "single_card":
            continue

        slug = _safe_str(payload.get("slug") or snapshot.id)
        name = _safe_str(payload.get("name") or slug)
        finish = _safe_str(payload.get("finish") or "Normal")
        condition = _safe_str(payload.get("condition") or "Near Mint (NM)")
        condition_tag = _condition_tag(condition)
        finish_label = _finish_label_for_title(finish)
        title = _truncate_title(f"Pokémon {name} {finish_label} ({condition_tag})", limit=60)

        base_price = round(_safe_float(payload.get("price_brl"), 0.0), 2)
        final_price, rule = _apply_price_rule(base_price, finish)
        quantity = max(0, _safe_int(payload.get("stock"), 0))
        image_url = _safe_str(payload.get("image_url"))
        language = (_safe_str(payload.get("language")) or "PT").upper()
        category = _safe_str(payload.get("category") or "Cartas avulsas")
        lot_code = _safe_str(payload.get("lot_id"))
        brand = _safe_str(payload.get("brand") or "Pokémon")
        universal_code = _safe_str(
            payload.get("gtin")
            or payload.get("ean")
            or payload.get("upc")
            or payload.get("barcode")
        )
        gtin_reason = "" if universal_code else "O produto não tem código cadastrado"
        description = _build_description(
            card_name=name,
            condition=condition,
            finish=finish,
            language=language,
            lot_code=lot_code,
        )

        rows.append(
            {
                "titulo": title,
                "descricao": description,
                "preco": final_price,
                "quantidade": quantity,
                "sku": slug,
                "foto_principal_url": image_url,
                "fotos_urls": image_url,
                "condicao": "Novo",
                "idioma": language,
                "categoria_interna": category,
                "slug": slug,
                "acabamento": finish,
                "preco_base_origem": base_price,
                "variacao_nome_carta": name,
                "codigo_universal_produto": universal_code,
                "motivo_gtin_vazio": gtin_reason,
                "unidade_venda": "Unidade",
                "tipo_anuncio": "Clássico",
                "marca": brand,
            }
        )
        counters["cards_total"] += 1
        counters[f"rule::{rule}"] += 1
        if quantity <= 0:
            counters["cards_stock_zero"] += 1
        if not image_url:
            counters["cards_without_image"] += 1

    rows.sort(key=lambda item: item["titulo"].lower())
    counters["db_cards_scanned"] = len(rows)
    counters["firestore_collection"] = settings.firestore_collection_products
    return rows, counters


def write_xlsx(rows: list[dict[str, Any]], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "cartas_ml"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_top = Alignment(vertical="top", wrap_text=True)

    ws.append(HEADERS)
    for col_idx, key in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 24
        if key in {"descricao", "foto_principal_url", "fotos_urls"}:
            ws.column_dimensions[cell.column_letter].width = 52
        if key in {"titulo"}:
            ws.column_dimensions[cell.column_letter].width = 44

    for row in rows:
        ws.append([row.get(key) for key in HEADERS])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = align_top

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def _build_header_alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for internal_key, aliases in TEMPLATE_HEADER_ALIASES.items():
        for alias in aliases:
            lookup[_normalize_header(alias)] = internal_key
    return lookup


def _build_header_aliases_by_internal() -> dict[str, set[str]]:
    output: dict[str, set[str]] = {}
    for internal_key, aliases in TEMPLATE_HEADER_ALIASES.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        output[internal_key] = {item for item in normalized_aliases if item}
    return output


def _resolve_internal_key_for_template_header(
    normalized_header: str,
    *,
    exact_lookup: dict[str, str],
    aliases_by_internal: dict[str, set[str]],
) -> str | None:
    if not normalized_header:
        return None

    exact = exact_lookup.get(normalized_header)
    if exact:
        return exact

    padded = f" {normalized_header} "
    for internal_key, aliases in aliases_by_internal.items():
        for alias in aliases:
            if (
                normalized_header.startswith(f"{alias} ")
                or normalized_header.endswith(f" {alias}")
                or f" {alias} " in padded
            ):
                return internal_key
    return None


def _resolve_template_data_start_row(sheet, header_row: int, mapping: dict[str, int]) -> int:
    scan_start = header_row + 1
    scan_end = min(sheet.max_row, header_row + 220)
    hint_terms = (
        "obrigatorio",
        "insira",
        "maximo",
        "identificar",
        "acessar",
        "selecione",
        "saiba",
        "revise",
        "este dado",
        "se a celula",
        "informacoes do produto",
    )

    for row_idx in range(scan_start, scan_end + 1):
        score = 0
        for key in ("preco", "quantidade", "titulo", "condicao"):
            col_idx = mapping.get(key)
            if not col_idx:
                continue
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            if isinstance(value, (int, float)):
                score += 2
                continue

            text = str(value).strip()
            if not text:
                continue
            if text.startswith("="):
                score += 2
                continue

            normalized_text = _normalize_header(text)
            if normalized_text and not any(term in normalized_text for term in hint_terms):
                score += 1

        if score >= 2:
            return row_idx

    return scan_start


def _resolve_template_target_sheet(workbook) -> tuple[Any, int, dict[str, int]]:
    alias_lookup = _build_header_alias_lookup()
    aliases_by_internal = _build_header_aliases_by_internal()
    best_score = -1
    best_sheet = None
    best_header_row = 0
    best_mapping: dict[str, int] = {}

    for sheet in workbook.worksheets:
        scan_limit = min(HEADER_SCAN_MAX_ROWS, max(1, sheet.max_row))
        for row_idx in range(1, scan_limit + 1):
            mapping: dict[str, int] = {}
            for col_idx in range(1, sheet.max_column + 1):
                normalized_header = _normalize_header(sheet.cell(row=row_idx, column=col_idx).value)
                if not normalized_header:
                    continue
                internal_key = _resolve_internal_key_for_template_header(
                    normalized_header,
                    exact_lookup=alias_lookup,
                    aliases_by_internal=aliases_by_internal,
                )
                if not internal_key or internal_key in mapping:
                    continue
                mapping[internal_key] = col_idx

            score = len(mapping)
            if score > best_score:
                best_score = score
                best_sheet = sheet
                best_header_row = row_idx
                best_mapping = mapping

    required = {"titulo", "preco", "quantidade"}
    if best_sheet is None or not required.issubset(best_mapping.keys()):
        raise RuntimeError(
            "Nao foi possivel localizar aba/cabecalho compativel no template do Mercado Livre. "
            "Use a planilha original baixada no Anunciador em Massa."
        )

    if "fotos_urls" not in best_mapping and "foto_principal_url" in best_mapping:
        best_mapping["fotos_urls"] = best_mapping["foto_principal_url"]
    if "foto_principal_url" not in best_mapping and "fotos_urls" in best_mapping:
        best_mapping["foto_principal_url"] = best_mapping["fotos_urls"]

    return best_sheet, best_header_row, best_mapping


def write_xlsx_using_template(
    rows: list[dict[str, Any]],
    *,
    template_xlsx: Path,
    output_xlsx: Path,
) -> tuple[str, int, dict[str, int]]:
    if not template_xlsx.exists():
        raise RuntimeError(f"Template nao encontrado: {template_xlsx}")

    workbook = load_workbook(template_xlsx, keep_vba=True)
    sheet, header_row, mapping = _resolve_template_target_sheet(workbook)
    data_start_row = _resolve_template_data_start_row(sheet, header_row, mapping)

    mapped_columns = sorted(set(mapping.values()))
    for row_idx in range(data_start_row, sheet.max_row + 1):
        for col_idx in mapped_columns:
            sheet.cell(row=row_idx, column=col_idx).value = None

    same_photo_column = (
        "foto_principal_url" in mapping
        and "fotos_urls" in mapping
        and mapping["foto_principal_url"] == mapping["fotos_urls"]
    )

    for offset, payload in enumerate(rows):
        row_idx = data_start_row + offset
        for key, col_idx in mapping.items():
            if same_photo_column and key == "fotos_urls":
                continue
            value = payload.get(key)
            if key == "preco":
                value = round(_safe_float(value, 0.0), 2)
            elif key == "quantidade":
                value = _safe_int(value, 0)
            elif value is not None:
                value = str(value)
            sheet.cell(row=row_idx, column=col_idx).value = value

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    return sheet.title, data_start_row, mapping


def _to_currency_brl(value: Any) -> str:
    amount = _safe_float(value, 0.0)
    text = f"{amount:,.2f}"
    text = text.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {text}"


def write_html_preview(rows: list[dict[str, Any]], html_path: Path, source_file: str) -> int:
    html_lines = [
        "<!doctype html>",
        "<html lang='pt-BR'>",
        "<head>",
        "  <meta charset='utf-8' />",
        "  <meta name='viewport' content='width=device-width, initial-scale=1' />",
        "  <title>Preview Mercado Livre - Cartas Legacy Cards</title>",
        "  <style>",
        "    body { font-family: Arial, sans-serif; margin: 20px; background: #f7f8fb; color: #12233d; }",
        "    h1 { margin: 0 0 8px; }",
        "    .meta { margin-bottom: 20px; color: #425b84; }",
        "    table { width: 100%; border-collapse: collapse; background: #fff; }",
        "    th, td { border: 1px solid #d3dced; padding: 8px; text-align: left; vertical-align: top; font-size: 13px; }",
        "    th { background: #1e3a8a; color: #fff; position: sticky; top: 0; }",
        "    img { width: 72px; height: 100px; object-fit: cover; border-radius: 8px; border: 1px solid #d3dced; }",
        "    .price { font-weight: 700; color: #0a5f2e; }",
        "    .rule { font-size: 12px; color: #4a5d7d; }",
        "    .desc { max-width: 440px; }",
        "  </style>",
        "</head>",
        "<body>",
        "  <h1>Preview Mercado Livre - Cartas Legacy Cards</h1>",
        f"  <div class='meta'>Gerado em {datetime.now(UTC).isoformat()} | Arquivo: {html.escape(source_file)}</div>",
        "  <table>",
        "    <thead><tr><th>Imagem</th><th>Titulo</th><th>Preco</th><th>Qtd</th><th>SKU</th><th>Acabamento</th><th>Descricao</th></tr></thead>",
        "    <tbody>",
    ]

    for row in rows:
        img = _safe_str(row.get("foto_principal_url"))
        titulo = _safe_str(row.get("titulo"))
        preco = _to_currency_brl(row.get("preco"))
        qtd = _safe_int(row.get("quantidade"), 0)
        sku = _safe_str(row.get("sku"))
        acabamento = _safe_str(row.get("acabamento"))
        descricao = _safe_str(row.get("descricao"))

        img_html = (
            f"<img src='{html.escape(img)}' alt='{html.escape(titulo)}' loading='lazy' />"
            if img
            else "<span>Sem imagem</span>"
        )
        html_lines.extend(
            [
                "      <tr>",
                f"        <td>{img_html}</td>",
                f"        <td>{html.escape(titulo)}</td>",
                f"        <td class='price'>{html.escape(preco)}</td>",
                f"        <td>{qtd}</td>",
                f"        <td>{html.escape(sku)}</td>",
                f"        <td>{html.escape(acabamento)}</td>",
                f"        <td class='desc'>{html.escape(descricao)}</td>",
                "      </tr>",
            ]
        )

    html_lines.extend(["    </tbody>", "  </table>", "</body>", "</html>"])
    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text("\n".join(html_lines), encoding="utf-8")
    return len(rows)


def main() -> None:
    args = parse_args()
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    output_dir = Path(args.output_dir).resolve()
    output_xlsx = output_dir / f"mercadolivre-cartas-legacy-{timestamp}.xlsx"
    output_html = output_dir / f"mercadolivre-cartas-legacy-{timestamp}.html"

    rows, counters = fetch_cards_rows(database=args.database)
    if not rows:
        raise SystemExit("Nenhuma carta encontrada para gerar planilha.")

    template_path = Path(args.template_xlsx).expanduser().resolve() if args.template_xlsx else None
    if template_path:
        sheet_name, data_start_row, mapping = write_xlsx_using_template(
            rows,
            template_xlsx=template_path,
            output_xlsx=output_xlsx,
        )
        print(
            "Template Mercado Livre preenchido:",
            f"aba={sheet_name}",
            f"data_start_row={data_start_row}",
            f"colunas_mapeadas={sorted(mapping.keys())}",
        )
    else:
        write_xlsx(rows, output_xlsx)
        print(
            "Arquivo gerado em formato livre (nao-template). "
            "Para upload no ML, prefira usar --template-xlsx com planilha original.",
        )

    html_count = write_html_preview(rows, output_html, output_xlsx.name)

    print("Arquivo XLSX gerado:", output_xlsx)
    print("Preview HTML gerado:", output_html)
    print("Total de cartas exportadas:", len(rows))
    print("Total de linhas no HTML:", html_count)
    print("Resumo de regras de preco:")
    for key, count in sorted(counters.items()):
        if key.startswith("rule::"):
            print(f"  - {key.replace('rule::', '')}: {count}")
    if counters.get("cards_without_image", 0):
        print("Aviso: cartas sem imagem:", counters["cards_without_image"])
    if counters.get("cards_stock_zero", 0):
        print("Aviso: cartas com estoque zero:", counters["cards_stock_zero"])

    if args.open_html:
        webbrowser.open(output_html.as_uri())
        print("Preview aberto no navegador.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
